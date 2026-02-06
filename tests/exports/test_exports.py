import json
import re
import unittest
from io import BytesIO

import pandas as pd
from openpyxl import load_workbook

from utils.exports.clinical_exports import (
    _count_mapping,
    _generate_csv_with_metadata,
    _generate_filename,
)
from utils.exports.search_json import export_search_json
from utils.exports.search_excel import generate_search_excel
from utils.exports.report_json import build_report_json


def _sample_search() -> dict:
    return {
        "id": "SEARCH-1",
        "name": "Main Search",
        "description": "Test search",
        "folder_id": "F-1",
        "dependencies": ["SEARCH-2"],
        "criteria_groups": [
            {
                "group_flags": {
                    "member_operator": "AND",
                    "action_if_true": "SELECT",
                    "action_if_false": "REJECT",
                },
                "criteria": [
                    {
                        "table": "EVENTS",
                        "column": "EVENT_CODE",
                        "flags": {
                            "logical_table_name": "EVENTS",
                            "display_name": "Clinical Codes",
                            "negation": False,
                        },
                        "value_sets": [
                            {
                                "code_system": "SNOMED_CONCEPT",
                                "code_value": "111",
                                "display_name": "Code One",
                                "valueSet_description": "VS 1",
                                "valueSet_guid": "VS-1",
                                "include_children": True,
                                "is_refset": False,
                            }
                        ],
                        "column_filters": [],
                        "linked_criteria": [],
                    }
                ],
                "population_criteria": [{"report_guid": "SEARCH-2"}],
            }
        ],
    }


class TestClinicalExportHelpers(unittest.TestCase):
    def test_count_mapping_and_filename_generation(self):
        df = pd.DataFrame(
            [
                {"Mapping Found": "Found"},
                {"Mapping Found": "Not Found"},
                {"Mapping Found": "found"},
            ]
        )
        total, matched, unmatched = _count_mapping(df)
        self.assertEqual((total, matched, unmatched), (3, 2, 1))

        filename = _generate_filename("Standalone Medications", "unique_codes", "matched")
        self.assertTrue(filename.startswith("standalone_medications_unique_matched_"))
        self.assertRegex(filename, r"\d{8}_\d{6}\.csv$")

    def test_generate_csv_with_metadata_includes_header_rows(self):
        df = pd.DataFrame([{"EMIS GUID": "A1", "SNOMED Code": "123"}])
        content = _generate_csv_with_metadata(df, "example.xml")
        lines = content.splitlines()

        self.assertTrue(lines[0].startswith("# Original XML File: example.xml"))
        self.assertTrue(lines[1].startswith("# Export Date/Time: "))
        self.assertEqual(lines[4], "EMIS GUID,SNOMED Code")
        self.assertEqual(lines[5], "A1,123")


class TestSearchAndReportExports(unittest.TestCase):
    def test_export_search_json_builds_expected_structure(self):
        search = _sample_search()
        parent = {"id": "SEARCH-2", "name": "Parent Population", "criteria_groups": []}
        all_searches = [search, parent]
        folders = [{"id": "F-1", "name": "Folder A", "parent_id": None}]
        id_to_name = {"SEARCH-1": "Main Search", "SEARCH-2": "Parent Population"}

        json_text = export_search_json("SEARCH-1", all_searches, folders, id_to_name)
        payload = json.loads(json_text)

        self.assertEqual(payload["metadata"]["search_id"], "SEARCH-1")
        self.assertEqual(payload["search"]["name"], "Main Search")
        self.assertEqual(len(payload["search"]["rules"]), 1)
        self.assertEqual(payload["search"]["rules"][0]["criteria"][0]["clinical_codes"][0]["code"], "111")

    def test_generate_search_excel_returns_workbook_bytes(self):
        search = _sample_search()
        parent = {"id": "SEARCH-2", "name": "Parent Population", "criteria_groups": []}
        all_searches = [search, parent]
        folders = [{"id": "F-1", "name": "Folder A", "parent_id": None}]
        id_to_name = {"SEARCH-1": "Main Search", "SEARCH-2": "Parent Population"}

        excel_bytes = generate_search_excel(search, all_searches, folders, id_to_name)
        self.assertIsInstance(excel_bytes, bytes)
        self.assertGreater(len(excel_bytes), 0)

        wb = load_workbook(BytesIO(excel_bytes))
        self.assertIn("Overview", wb.sheetnames)
        self.assertIn("Rule 1 Logic", wb.sheetnames)
        self.assertIn("Rule 1 Codes", wb.sheetnames)

    def test_build_report_json_returns_filename_and_serialized_payload(self):
        report = {
            "id": "REPORT-1",
            "name": "List Report",
            "type": "list",
            "type_label": "List Report",
            "description": "Report description",
            "folder_path": ["Root", "Subfolder"],
            "author": "Tester",
            "creation_time": "2026-02-03T10:00:00",
            "column_groups": [],
            "report_criteria": [],
        }

        filename, payload_text = build_report_json(report)
        payload = json.loads(payload_text)

        self.assertTrue(filename.endswith(".json"))
        self.assertTrue(re.match(r"List_Report_list_\d{8}_\d{6}\.json$", filename))
        self.assertEqual(payload["export_metadata"]["report_id"], "REPORT-1")
        self.assertEqual(payload["report"]["overview"]["report_name"], "List Report")


if __name__ == "__main__":
    unittest.main()
