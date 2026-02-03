"""
Excel export for reports.
"""

from __future__ import annotations

import io
from typing import Any, Dict, Tuple, Optional
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from .report_export_common import build_report_filename, build_report_overview, build_report_tables


HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _apply_header_style(ws) -> None:
    if ws.max_row < 1:
        return
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT


def _auto_size_columns(ws, max_width: int = 60) -> None:
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            if cell.value is None:
                continue
            value_length = len(str(cell.value))
            if value_length > max_length:
                max_length = value_length
        if max_length:
            ws.column_dimensions[column_letter].width = min(max_length + 2, max_width)


def build_report_excel(
    report: Dict[str, Any],
    id_to_name: Optional[Dict[str, str]] = None,
) -> Tuple[str, bytes]:
    output = io.BytesIO()

    overview_rows = build_report_overview(report, id_to_name)
    overview_df = pd.DataFrame(overview_rows, columns=["Property", "Value"])
    tables = build_report_tables(report, id_to_name)

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        overview_df.to_excel(writer, sheet_name="Report_Overview", index=False)

        for section, rows in tables.items():
            if section == "Overview":
                continue
            df = pd.DataFrame(rows)
            if df.empty:
                continue
            sheet_name = _safe_sheet_name(section)
            df.to_excel(writer, sheet_name=sheet_name, index=False)

        for ws in writer.book.worksheets:
            _apply_header_style(ws)
            _auto_size_columns(ws)

    output.seek(0)
    filename = build_report_filename(report, "report", "xlsx")
    return filename, output.getvalue()

def _safe_sheet_name(name: str) -> str:
    cleaned = "".join(ch for ch in name if ch.isalnum() or ch in (" ", "_")).strip()
    if not cleaned:
        return "Sheet"
    return cleaned[:31]
