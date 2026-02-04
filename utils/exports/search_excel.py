"""
Excel export functionality for searches.
Generates comprehensive Excel workbooks with Overview, Rule Logic, and Rule Codes tabs.
"""

from io import BytesIO
from typing import Dict, List, Any
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from .search_data_provider import (
    get_search_by_id,
    get_search_dependencies,
    get_search_folder_path,
    extract_rule_criteria,
    extract_criterion_codes,
    extract_date_restrictions,
    extract_event_restrictions,
    extract_value_restrictions,
    extract_parameters,
    get_population_references,
    format_date_restriction,
    format_event_restriction,
    format_value_restriction,
    format_rule_action,
)
from ..metadata.emisinternal_describer import describe_emisinternal_filter
from ..metadata.value_set_resolver import resolve_value_sets


# Styling constants
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
ALT_ROW_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
NESTED_FILL = PatternFill(start_color="E7F1FF", end_color="E7F1FF", fill_type="solid")
DIVIDER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
DIVIDER_FONT = Font(bold=True)
CENTRE_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGNMENT = Alignment(horizontal="left", vertical="top", wrap_text=True)


EMOJI_RANGES = (
    (0x2600, 0x26FF),    # misc symbols
    (0x2700, 0x27BF),    # dingbats
    (0x1F1E6, 0x1F1FF),  # flags
    (0x1F300, 0x1F5FF),  # symbols & pictographs
    (0x1F600, 0x1F64F),  # emoticons
    (0x1F680, 0x1F6FF),  # transport & map symbols
    (0x1F900, 0x1FAFF),  # supplemental symbols and pictographs
)
EMOJI_CODEPOINTS = {
    0xFE0F,  # variation selector
    0x200D,  # zero width joiner
}


def _apply_header_style(ws, row_num: int):
    """Apply header styling to a row."""
    for cell in ws[row_num]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTRE_ALIGNMENT


def _strip_emojis(text: str) -> str:
    """Remove emojis and other pictographs from text."""
    cleaned_chars: List[str] = []
    for ch in str(text):
        codepoint = ord(ch)
        if codepoint in EMOJI_CODEPOINTS:
            continue
        if any(start <= codepoint <= end for start, end in EMOJI_RANGES):
            continue
        cleaned_chars.append(ch)
    return "".join(cleaned_chars).strip()


def _auto_size_columns(ws, max_width: int = 60):
    """Auto-size columns based on content with max width limit."""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    max_length = max(max_length, cell_length)
            except:
                pass

        adjusted_width = min(max_length + 2, max_width)
        ws.column_dimensions[column_letter].width = adjusted_width


def _add_divider_row(ws, row_num: int, operator: str, col_count: int):
    """Add operator divider row with styling."""
    divider_text = f"─── {operator.upper()} ───"

    # Merge cells across all columns
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=col_count)

    cell = ws.cell(row=row_num, column=1, value=divider_text)
    cell.fill = DIVIDER_FILL
    cell.font = DIVIDER_FONT
    cell.alignment = CENTRE_ALIGNMENT


def _generate_overview_tab(
    wb: Workbook,
    search: Dict,
    all_searches: List[Dict],
    folders: List[Dict],
    id_to_name: Dict
) -> None:
    """Generate the Overview tab with search metadata."""
    ws = wb.create_sheet("Overview", 0)

    # Headers
    headers = ["Field", "Value"]
    ws.append(headers)
    _apply_header_style(ws, 1)

    # Get dependencies
    search_id = search.get("id")
    dependencies = get_search_dependencies(search_id, all_searches)

    # Resolve dependant names
    dependant_names = []
    for dep_id in dependencies["dependants"]:
        dep_search = get_search_by_id(dep_id, all_searches)
        if dep_search:
            dependant_names.append(_strip_emojis(dep_search.get("name", dep_id)))

    # Resolve parent names
    parent_names = []
    for parent_id in dependencies["parents"]:
        parent_search = get_search_by_id(parent_id, all_searches)
        if parent_search:
            parent_names.append(_strip_emojis(parent_search.get("name", parent_id)))

    # Extract parameters
    parameters = extract_parameters(search)
    param_details = []
    for param in parameters:
        param_details.append(f"{param['name']} ({param['data_type']}, {param['scope']})")

    # Count total codes across all rules
    total_codes = 0
    total_criteria = 0
    for group in search.get("criteria_groups", []):
        criteria_list = extract_rule_criteria(group)
        total_criteria += len(criteria_list)

        for crit_info in criteria_list:
            codes = extract_criterion_codes(
                crit_info["criterion"],
                crit_info["criterion_number"]
            )
            total_codes += len(codes)

    # Get export timestamp
    from datetime import datetime
    export_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Build rows
    rows = [
        ["Export Date/Time", export_time],
        ["Search ID", search_id],
        ["Search Name", search.get("name", "")],
        ["Description", search.get("description", "")],
        ["Folder Path", get_search_folder_path(search_id, folders, all_searches)],
        ["Total Rules", len(search.get("criteria_groups", []))],
        ["Total Criteria", total_criteria],
        ["Total Clinical Codes", total_codes],
        ["Has Parameters", "Yes" if parameters else "No"],
        ["Parameter Details", "\n".join(param_details) if param_details else "N/A"],
        ["Parent Populations", "\n".join(parent_names) if parent_names else "None"],
        ["Used By (Dependants)", "\n".join(dependant_names) if dependant_names else "None"],
        ["", ""],  # Blank row
        ["Generated by", "ClinXML™ EMIS XML Converter (https://clinxml.streamlit.app)"],
    ]

    for row_data in rows:
        ws.append(row_data)

    # Apply alternating row colours
    for idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        if idx % 2 == 0:
            for cell in row:
                cell.fill = ALT_ROW_FILL
        for cell in row:
            cell.alignment = LEFT_ALIGNMENT

    _auto_size_columns(ws)
    ws.freeze_panes = "A2"


def _generate_rule_logic_tab(
    wb: Workbook,
    rule_number: int,
    group: Dict,
    search: Dict,
    all_searches: List[Dict],
    id_to_name: Dict
) -> None:
    """Generate Rule Logic tab for a specific rule with vertical layout."""
    ws = wb.create_sheet(f"Rule {rule_number} Logic")

    group_flags = group.get("group_flags", {})
    operator = group_flags.get("operator") or group.get("operator") or group_flags.get("member_operator") or "AND"
    operator = operator.upper()

    # Store rule actions
    action_if_true_raw = group_flags.get("action_if_true") or "SELECT"
    action_if_false_raw = group_flags.get("action_if_false") or "REJECT"

    # Add SCORE threshold if applicable
    if operator == "SCORE":
        score_range = group_flags.get("score_range", {})
        if score_range:
            min_score = score_range.get("min_score", "")
            operator_sym = score_range.get("operator", "GTEQ")

            op_text = {
                "GTEQ": "greater than or equal to",
                "GT": "greater than",
                "LTEQ": "less than or equal to",
                "LT": "less than",
                "EQ": "equal to"
            }.get(operator_sym, operator_sym)

            ws.append([f"SCORE Rule: Only include patients who occur in {op_text} {min_score} of the following:"])
            ws.append([])

    criteria_list = extract_rule_criteria(group)

    # Vertical layout - render each criterion
    for idx, crit_info in enumerate(criteria_list):
        _render_criterion_vertical(ws, crit_info, operator, id_to_name, all_searches, group)

        # Add operator divider between same-level criteria
        if idx < len(criteria_list) - 1:
            next_crit = criteria_list[idx + 1]
            if next_crit["nesting_level"] == crit_info["nesting_level"]:
                ws.append([])
                ws.append([f"--- {operator} ---"])
                ws.append([])

    # If no criteria but there are population references, show them
    if not criteria_list:
        pop_refs = get_population_references(group, id_to_name, all_searches)
        if pop_refs:
            for idx, ref in enumerate(pop_refs):
                search_name = _strip_emojis(ref["search_name"])
                score_weight = ref.get("score_weight", "N/A")

                ws.append([f"Using search: {search_name}"])
                ws.append([f"  Score: {score_weight}"])

                if idx < len(pop_refs) - 1:
                    ws.append([])
                    ws.append([f"--- {operator} ---"])
                    ws.append([])

    # Add rule actions at the bottom
    ws.append([])
    ws.append([])
    ws.append(["If rule passed:", format_rule_action(action_if_true_raw)])
    ws.append(["If rule failed:", format_rule_action(action_if_false_raw)])

    _auto_size_columns(ws)


def _render_criterion_vertical(
    ws,
    crit_info: Dict,
    operator: str,
    id_to_name: Dict,
    all_searches: List[Dict],
    group: Dict
) -> None:
    """Render a single criterion in vertical layout."""
    criterion = crit_info["criterion"]
    flags = criterion.get("flags", {})
    crit_num = crit_info["criterion_number"]
    nesting_level = crit_info["nesting_level"]
    relationship = crit_info.get("relationship")

    indent = "  " * nesting_level

    # Header
    criterion_name = flags.get("display_name") or "Clinical Codes"
    if relationship:
        ws.append([f"{indent}Criterion {crit_num}: (Linked Feature)"])

        # Show relationship
        from ..parsing.node_parsers.linked_criteria_parser import get_temporal_relationship_description
        parent_col = relationship.get("parent_column_display_name") or relationship.get("parent_column") or "parent field"
        child_col = relationship.get("child_column_display_name") or relationship.get("child_column") or "child field"
        temporal_desc = get_temporal_relationship_description(relationship)

        if temporal_desc:
            ws.append([f"{indent}  Relationship: {parent_col} is {temporal_desc} {child_col}"])
        else:
            ws.append([f"{indent}  Relationship: {parent_col} relates to {child_col}"])
    else:
        ws.append([f"{indent}Criterion {crit_num}: {criterion_name}"])

    # Metadata
    table = flags.get("logical_table_name") or criterion.get("table", "Unknown")
    action = "Exclude" if flags.get("negation") else "Include"
    ws.append([f"{indent}  Table: {table}"])
    ws.append([f"{indent}  Action: {action}"])

    score_weight = flags.get("score_weightage")
    if score_weight:
        ws.append([f"{indent}  Score Weight: {score_weight}"])

    # Clinical Codes and Library Items summary
    codes = extract_criterion_codes(criterion, crit_num)

    # Check for library items
    library_items = []
    for vs in resolve_value_sets(criterion):
        code_system = (vs.get("code_system") or "").upper()
        if code_system == "LIBRARY_ITEM":
            lib_id = vs.get("valueSet_guid") or vs.get("id")
            if not lib_id:
                # Try to get from codes
                vs_codes = vs.get("codes") or []
                if vs_codes:
                    lib_id = vs_codes[0].get("EMIS GUID") or vs_codes[0].get("emis_guid") or vs_codes[0].get("code_value")
            if lib_id:
                library_items.append(lib_id)

    if codes or library_items:
        ws.append([])
        ws.append([f"{indent}  Clinical Codes:"])

        # Show library items
        for lib_id in library_items:
            ws.append([f"{indent}    - Library Item: {lib_id}"])

        # Show code count if there are non-library codes
        if codes:
            code_word = "code" if len(codes) == 1 else "codes"
            ws.append([f"{indent}    - {action} {len(codes)} specified clinical {code_word}"])

    # Filters section
    date_restr = extract_date_restrictions(criterion)
    event_restr = extract_event_restrictions(criterion)
    value_restr = extract_value_restrictions(criterion)

    # EMISINTERNAL filters - get column display names from column_filters
    emisinternal_filters = []
    seen_filters = set()

    # Build map of column name to display name from column_filters
    column_display_map = {}
    for cf in criterion.get("column_filters", []):
        col_name = (cf.get("column") or cf.get("column_name") or "").strip().upper()
        col_display = cf.get("column_display") or cf.get("display_name") or cf.get("column_name") or ""
        if col_name:
            column_display_map[col_name] = col_display

    # Use enriched entries from flags (these are processed by pattern plugins)
    emisinternal_entries = flags.get("emisinternal_entries") or []
    if emisinternal_entries:
        for entry in emisinternal_entries:
            values = entry.get("values") or []
            if values:
                raw_column = (entry.get("column") or "").strip().upper()
                # Get display name from column_filters, fallback to raw column
                column_display = column_display_map.get(raw_column) or entry.get("column") or ""
                vs_like = [{"values": values}]
                in_not_in = str(entry.get("in_not_in", "")).upper()
                desc = describe_emisinternal_filter(column_display, vs_like, in_not_in=in_not_in)
                if entry.get("has_all_values"):
                    desc += " (all values)"

                # Deduplicate
                if desc not in seen_filters:
                    emisinternal_filters.append(desc)
                    seen_filters.add(desc)
    else:
        # Fallback to value_sets if no enriched entries
        for vs in resolve_value_sets(criterion):
            code_system = (vs.get("code_system") or "").upper()
            if code_system == "EMISINTERNAL":
                column_name = vs.get("column_display_name") or vs.get("column_name") or vs.get("column", "")
                in_not_in = vs.get("in_not_in", "IN")
                vs_like = [{"values": vs.get("values", [])}]
                desc = describe_emisinternal_filter(column_name, vs_like, in_not_in)

                # Deduplicate
                if desc not in seen_filters:
                    emisinternal_filters.append(desc)
                    seen_filters.add(desc)

    if date_restr or event_restr or value_restr or emisinternal_filters:
        ws.append([])
        ws.append([f"{indent}  Filters:"])

        for dr in date_restr:
            ws.append([f"{indent}    - {format_date_restriction(dr)}"])
        for er in event_restr:
            ws.append([f"{indent}    - {format_event_restriction(er)}"])
        for vr in value_restr:
            ws.append([f"{indent}    - {format_value_restriction(vr)}"])
        for ef in emisinternal_filters:
            ws.append([f"{indent}    - {ef}"])

    # Demographics
    has_demographics = (
        flags.get("has_demographic_filters") or
        flags.get("is_patient_demographics") or
        flags.get("demographics_type")
    )

    if has_demographics:
        ws.append([])
        ws.append([f"{indent}  Demographics:"])

        # LSOA demographics
        if flags.get("demographics_type") == "LSOA":
            lsoa_codes = flags.get("consolidated_lsoa_codes", [])
            count = flags.get("consolidated_count", len(lsoa_codes))
            column_display = flags.get("column_display_name") or "Lower Layer Area"

            if lsoa_codes:
                ws.append([f"{indent}    - {column_display} ({len(lsoa_codes)} areas):"])
                # Show all codes
                for code in lsoa_codes:
                    ws.append([f"{indent}      {code}"])
        # Standard demographics (age/sex/status)
        elif flags.get("has_demographic_filters"):
            if flags.get("age_min") is not None:
                ws.append([f"{indent}    - Age >= {flags['age_min']}"])
            if flags.get("age_max") is not None:
                ws.append([f"{indent}    - Age <= {flags['age_max']}"])
            if flags.get("sex_filter"):
                ws.append([f"{indent}    - Sex: {flags['sex_filter']}"])
            if flags.get("patient_status"):
                ws.append([f"{indent}    - Status: {flags['patient_status']}"])

    # Population references (for top-level only)
    if not nesting_level:
        pop_refs = get_population_references(group, id_to_name, all_searches)
        if pop_refs:
            ws.append([])
            ws.append([f"{indent}  Using Searches:"])
            for ref in pop_refs:
                search_name = _strip_emojis(ref["search_name"])
                score = ref.get("score_weight")
                if score:
                    ws.append([f"{indent}    - {search_name} (Score: {score})"])
                else:
                    ws.append([f"{indent}    - {search_name}"])

    # Parameters with scope (deduplicated)
    params = criterion.get("parameters") or flags.get("parameter_names") or []
    if params:
        # Determine default scope
        scope_bits = []
        if flags.get("has_global_parameters"):
            scope_bits.append("Global")
        if flags.get("has_local_parameters"):
            scope_bits.append("Local")
        default_scope = " & ".join(scope_bits) if scope_bits else "Unknown"

        # Build formatted parameter strings with scope
        param_strings = []
        seen = set()
        for p in params:
            if isinstance(p, dict):
                name = p.get("name") or str(p)
                # Check if this parameter has its own scope info
                p_scope_bits = []
                if p.get("allowGlobal") or p.get("is_global"):
                    p_scope_bits.append("Global")
                if p.get("allowLocal") or p.get("is_local"):
                    p_scope_bits.append("Local")
                p_scope = " & ".join(p_scope_bits) if p_scope_bits else default_scope
                param_str = f"[Name: {name}, Scope: {p_scope}]"
            else:
                name = str(p)
                param_str = f"[Name: {name}, Scope: {default_scope}]"

            # Deduplicate
            if param_str not in seen:
                param_strings.append(param_str)
                seen.add(param_str)

        ws.append([])
        ws.append([f"{indent}  Parameters: {', '.join(param_strings)}"])

    # Notes
    notes = flags.get("exception_code")
    if notes:
        ws.append([])
        ws.append([f"{indent}  Notes: {notes}"])

    ws.append([])


def _generate_rule_codes_tab(
    wb: Workbook,
    rule_number: int,
    group: Dict,
    search: Dict
) -> None:
    """Generate Rule Codes tab for a specific rule."""
    ws = wb.create_sheet(f"Rule {rule_number} Codes")

    # Headers
    headers = [
        "Criterion #",
        "Parent Criterion",
        "Code System",
        "Code",
        "Term",
        "ValueSet Name",
        "ValueSet GUID",
        "Is Refset",
        "Include Children",
        "Pseudo Refset Member",
        "Negation"
    ]
    ws.append(headers)
    _apply_header_style(ws, 1)

    # Extract all criteria and their codes
    criteria_list = extract_rule_criteria(group)

    all_codes = []
    for crit_info in criteria_list:
        criterion = crit_info["criterion"]
        crit_num = crit_info["criterion_number"]
        parent_num = crit_info["parent_number"]

        codes = extract_criterion_codes(criterion, crit_num)

        for code in codes:
            all_codes.append({
                "criterion_number": crit_num,
                "parent_criterion": parent_num if parent_num else "N/A",
                "code_system": code["code_system"],
                "code": code["code"],
                "term": code["term"],
                "valueset_name": code["valueset_name"],
                "valueset_guid": code["valueset_guid"],
                "is_refset": "Yes" if code["is_refset"] else "No",
                "include_children": "Yes" if code["include_children"] else "No",
                "pseudo_refset_member": "Yes" if code.get("pseudo_refset_member") else "No",
                "negation": "Exclude" if code["negation"] else "Include"
            })

    # Sort by criterion number then code
    all_codes.sort(key=lambda x: (x["criterion_number"], x["code_system"], x["code"]))

    # Add rows
    for idx, code_data in enumerate(all_codes, start=2):
        row = [
            code_data["criterion_number"],
            code_data["parent_criterion"],
            code_data["code_system"],
            code_data["code"],
            code_data["term"],
            code_data["valueset_name"],
            code_data["valueset_guid"],
            code_data["is_refset"],
            code_data["include_children"],
            code_data["pseudo_refset_member"],
            code_data["negation"]
        ]
        ws.append(row)

        # Apply alternating colours
        if idx % 2 == 0:
            for cell in ws[idx]:
                cell.fill = ALT_ROW_FILL

        # Apply alignment
        for cell in ws[idx]:
            cell.alignment = LEFT_ALIGNMENT

    _auto_size_columns(ws)
    ws.freeze_panes = "A2"


def generate_search_excel(
    current_search: Dict,
    all_searches: List[Dict],
    folders: List[Dict],
    id_to_name: Dict
) -> bytes:
    """
    Generate complete Excel workbook for a search.

    Returns bytes of the Excel file ready for download.
    """
    wb = Workbook()

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Generate Overview tab
    _generate_overview_tab(wb, current_search, all_searches, folders, id_to_name)

    # Generate Rule Logic and Rule Codes tabs for each rule
    for rule_idx, group in enumerate(current_search.get("criteria_groups", []), start=1):
        _generate_rule_logic_tab(wb, rule_idx, group, current_search, all_searches, id_to_name)
        _generate_rule_codes_tab(wb, rule_idx, group, current_search)

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output.getvalue()
